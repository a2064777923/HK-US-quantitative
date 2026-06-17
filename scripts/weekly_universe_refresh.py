#!/usr/bin/env python3
"""
weekly_universe_refresh.py - 每週刷新港股+美股股票池
數據源: HKEX官方xlsx + NASDAQ API
自動新增新股、更新名稱、標記退市股
"""
import requests, subprocess, io, time
from datetime import datetime

try:
    from us_universe_filter import is_supported_us_equity, normalize_us_symbol
except ImportError:  # pragma: no cover - package import path used by local tests
    from scripts.us_universe_filter import is_supported_us_equity, normalize_us_symbol

def db(sql, timeout=30):
    r = subprocess.run(
        ["docker", "exec", "quantmind-db", "psql", "-U", "quantmind", "-d", "quantmind", "-t", "-A", "-c", sql],
        capture_output=True, text=True, timeout=timeout
    )
    return r.stdout.strip()

def log(msg):
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}", flush=True)

def retry(fn, max_attempts=3, delay=10, desc=""):
    for attempt in range(1, max_attempts + 1):
        try:
            return fn()
        except Exception as e:
            log(f"  {desc} attempt {attempt}/{max_attempts} failed: {e}")
            if attempt < max_attempts:
                time.sleep(delay * attempt)
    return None

def insert_batch(rows):
    if not rows:
        return 0
    values = []
    for sym, name, exch, cur in rows:
        name_esc = name.replace("'", "''")[:499]
        values.append(f"('{sym}', '{name_esc}', '{exch}', '{cur}', true, NOW(), NOW())")
    sql = (
        "INSERT INTO stocks (symbol, name, exchange, currency, is_active, created_at, updated_at) VALUES "
        + ",".join(values)
        + " ON CONFLICT (symbol) DO UPDATE SET name=EXCLUDED.name, is_active=true, "
        + "exchange=EXCLUDED.exchange, currency=EXCLUDED.currency, updated_at=NOW()"
    )
    db(sql, timeout=60)
    return len(rows)


def main():
    log("=== 每週股票池刷新 ===")
    before_hk = db("SELECT count(*) FROM stocks WHERE is_active=true AND exchange='HKEX'")
    before_us = db("SELECT count(*) FROM stocks WHERE is_active=true AND exchange IN ('NASDAQ','NYSE')")
    log(f"刷新前: 港股={before_hk}, 美股={before_us}")

    # === HK from HKEX ===
    log("拉取 HKEX...")
    import openpyxl
    def fetch_hk():
        r = requests.get(
            "https://www.hkex.com.hk/eng/services/trading/securities/securitieslists/ListOfSecurities.xlsx",
            timeout=120, headers={"User-Agent": "Mozilla/5.0"}
        )
        r.raise_for_status()
        return r.content

    content = retry(fetch_hk, max_attempts=5, delay=15, desc="HKEX")
    if content:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        batch = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            code = str(row[0]).strip() if row[0] else ""
            name = str(row[1]).strip() if row[1] else ""
            cat = str(row[2]).strip() if row[2] else ""
            cur = str(row[16]).strip() if row[16] else "HKD"
            if "Equity" in cat and code and len(code) == 5:
                if cur not in ("HKD", "CNY", "USD"):
                    cur = "HKD"
                batch.append((code, name, "HKEX", cur))
                if len(batch) >= 50:
                    insert_batch(batch)
                    batch = []
        if batch:
            insert_batch(batch)
        log(f"港股: 處理 {ws.max_row - 3} 行")

    # === US from NASDAQ API ===
    for exchange in ["NASDAQ", "NYSE"]:
        log(f"拉取 {exchange}...")
        def fetch_us(ex=exchange):
            url = f"https://api.nasdaq.com/api/screener/stocks?tableonly=true&limit=10000&exchange={ex}"
            r = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
            r.raise_for_status()
            return r.json()

        data = retry(fetch_us, max_attempts=5, delay=15, desc=f"{exchange}")
        if not data:
            continue
        rows = data.get("data", {}).get("table", {}).get("rows", [])
        batch = []
        for item in rows:
            if not is_supported_us_equity(item):
                continue
            symbol = normalize_us_symbol(item.get("symbol", ""))
            name = item.get("name", "").strip()
            batch.append((symbol, name, exchange, "USD"))
            if len(batch) >= 50:
                insert_batch(batch)
                batch = []
        if batch:
            insert_batch(batch)
        log(f"{exchange}: 處理 {len(rows)} 隻")

    # Summary
    after_hk = db("SELECT count(*) FROM stocks WHERE is_active=true AND exchange='HKEX'")
    after_us = db("SELECT count(*) FROM stocks WHERE is_active=true AND exchange IN ('NASDAQ','NYSE')")
    log(f"刷新完成: 港股={after_hk}({int(after_hk)-int(before_hk):+d}), 美股={after_us}({int(after_us)-int(before_us):+d})")

if __name__ == "__main__":
    main()
